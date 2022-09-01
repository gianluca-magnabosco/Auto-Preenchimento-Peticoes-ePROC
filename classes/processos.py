from openpyxl import load_workbook


class Processos():

    processos = []
    clientes = []
    adversas = []
    cidades = []
    numero_processos = []


    def getData(self):
        self.wb = load_workbook("arquivos/processos.xlsx")
        self.ws = self.wb.active

        for i in range(1, self.ws.max_row + 1):
            self.processos.append(self.ws.cell(row = i, column = 1).value)
            self.clientes.append(self.ws.cell(row = i, column = 2).value)
            self.adversas.append(self.ws.cell(row = i, column = 3).value)
            self.cidades.append(self.ws.cell(row = i, column = 4).value)

        for i in range(self.ws.max_row):
            self.numero_processos.append(self.processos[i][:25])
