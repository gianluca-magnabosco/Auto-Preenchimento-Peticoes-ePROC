from openpyxl import load_workbook
import pyexcel as p
import pyexcel_xls
import pyexcel_xlsx
import pandas as pd
import glob
import os


class FormatExcel():

    def formatFiles(self):
        self.getFiles()
        self.formatSheet("sc")
        self.formatSheet("pr")
        self.mergeSheets()
        self.orderMergedSheet()
    
    
    def getFiles(self):
        self.file_name = glob.glob("arquivos/*.xls")
        self.convertFileToXlsx("sc")
        self.convertFileToXlsx("pr")


    def convertFileToXlsx(self, estado):
        file_name = self.file_name[0]
        p.save_book_as(file_name = file_name, dest_file_name = f"arquivos/processos_{estado}.xlsx")
        os.remove(file_name)
        self.file_name.pop(0)


    def formatSheet(self, estado):
        wb = load_workbook(f"arquivos/processos_{estado}.xlsx")
        ws = wb.active

        if estado == "sc":
            self.scMaxRows = ws.max_row-2

        for _ in range(2):
            ws.delete_rows(1)

        ws.delete_cols(2)

        for _ in range(4):
            ws.delete_cols(5)

        wb.save(f"arquivos/processos_{estado}_formatado.xlsx")
        os.remove(f"arquivos/processos_{estado}.xlsx")


    def mergeSheets(self):
        scDataFrame = pd.DataFrame()
        scDataFrame = pd.read_excel("arquivos/processos_sc_formatado.xlsx", header=None)
        prDataFrame = pd.DataFrame()
        prDataFrame = pd.read_excel("arquivos/processos_pr_formatado.xlsx", header=None)

        with pd.ExcelWriter("arquivos/merged.xlsx", mode = "w") as writer:  
            scDataFrame.to_excel(writer, index = False, header = False, startrow = 0)
            prDataFrame.to_excel(writer, index = False, header = False, startrow = self.scMaxRows)
        
        os.remove("arquivos/processos_sc_formatado.xlsx")
        os.remove("arquivos/processos_pr_formatado.xlsx")


    def orderMergedSheet(self):
        mergedExcelFile = pd.ExcelFile("arquivos/merged.xlsx")
        mergedDataFrame = mergedExcelFile.parse("Sheet1", header = None)
        mergedDataFrame = mergedDataFrame.sort_values(mergedDataFrame.columns[0])

        with pd.ExcelWriter("arquivos/merged_and_ordered.xlsx", mode = "w") as writer:
            mergedDataFrame.to_excel(writer, index = False, header = False)

        finalExcelFile = pd.read_excel("arquivos/merged_and_ordered.xlsx", header = None)
        finalExcelFile = finalExcelFile.drop_duplicates(keep = "first")
        finalExcelFile.to_excel("arquivos/processos.xlsx", index = False, header = False)

        mergedExcelFile.close()
        os.remove("arquivos/merged.xlsx")
        os.remove("arquivos/merged_and_ordered.xlsx")



def main():
    format = FormatExcel()
    format.formatFiles()


if __name__ == "__main__":
    main()
