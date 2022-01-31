from openpyxl import load_workbook
import pyexcel as p
import pyexcel_xls
import pyexcel_xlsx
import os
import datetime
from datetime import date
import pandas as pd


# load file, sheet
wb = load_workbook('processos_sc.xlsx')
ws = wb.active

max_rows = ws.max_row-2

# formatting excel sheet
for i in range(0,2):
    ws.delete_rows(1)


ws.delete_cols(2)
for i in range(0,4):
    ws.delete_cols(5)

wb.save('processos_sc_formatado.xlsx')


#p.save_book_as(file_name= 'processos_pr.xls', dest_file_name='processos_pr.xlsx')

# load file, sheet
wb = load_workbook('processos_pr.xlsx')
ws = wb.active


# formatting excel sheet
for i in range(0,2):
    ws.delete_rows(1)

ws.delete_cols(2)
for i in range(0,4):
    ws.delete_cols(5)

wb.save('processos_pr_formatado.xlsx')


df1 = pd.DataFrame()
df1 = pd.read_excel('processos_sc_formatado.xlsx', header=None)
df2 = pd.DataFrame()
df2 = pd.read_excel('processos_pr_formatado.xlsx', header=None)

with pd.ExcelWriter('output.xlsx', mode='w') as writer:  
    df1.to_excel(writer, index=False,header=False, startrow=0)
    df2.to_excel(writer, index=False, header=False, startrow=max_rows)

xl = pd.ExcelFile("output.xlsx")
df = xl.parse("Sheet1", header=None)
df = df.sort_values(df.columns[0])

writer = pd.ExcelWriter('output_final.xlsx')
df.to_excel(writer, index=False, header=False)
writer.save()



data = pd.read_excel('output_final.xlsx', header=None)
final_excel_file = data.drop_duplicates(keep="first")
final_excel_file.to_excel('processos_final_ordenado.xlsx', index=False, header=False)